from openpyxl import Workbook
wb = Workbook()    # 새 워크북 생성 
ws = wb.create_sheet() # 새로운 sheet 를 기본 이름으로 생성 
ws.title = "MySheet"  # sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff"  # 탭 색상 변경 

ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 sheet 생성 
ws2 = wb.create_sheet("NewSheet",2) # 2번째 index 에 sheet 생성 

new_ws = wb["NewSheet"] # Dict 형태로 sheet 에 접근 

print(wb.sheetnames)  # sheet 명 출력 

# sheet 복사 

new_ws["A1"]="Test"    # A1 셀에 데이터 입력 
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")