# To add a new cell, type '# %%'
# To add a new markdown cell, type '# %% [markdown]'
# %% [markdown]
# # Quiz
# - 출석 10
# - 퀴즈1 10
# - 퀴즈2 10 
# - 중간고사 20
# - 기말고사 30
# - 프로젝트 20 
# 
# - 학번, 출석, 퀴즈1, 퀴즈2, 중간고사, 기말고사, 프로젝트, 총점, 학점
# - 퀴즈2 는 모두 만점 처리 
# - 90~ A, 80~ B, 70~ C, 나머지 D, 출석 5 미만 총점에 관계없이 F 

# %%
from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active


# %%
ws.append(["학번","출석","퀴즈1","퀴즈2","중간고사","기말고사","프로젝트","총점","학점"])

# %% [markdown]
# ## 데이터 넣기 

# %%
for i in range(1,11):
    ws.append([i,randint(3,10),randint(5,10),randint(5,10),randint(10,20),randint(15,30),randint(10,20)])

# %% [markdown]
# ## 퀴즈2 점수 수정  ( 모두 10점 만점 으로 갱신 )

# %%
row_range = ws[1:ws.max_row]
print("Before fix 퀴즈2")
for row in row_range:
    print(row[3].value)

row_range = ws[2:ws.max_row]
for row in row_range:
    row[3].value = 10

print("After fix 퀴즈2")
row_range = ws[1:ws.max_row]
for row in row_range:
    print(row[3].value)

## Teacher sol. 
# for inx, cell in enumerate(ws["D"]):
#     if idx ==0:   # 제목인 경우 skip 
#         continue
#     cell.value = 10 

# %% [markdown]
# ## 총점 계산 

# %%
row_range = ws[2:ws.max_row]

for row in row_range:
    sum = 0 
    for i in range(1,7):
        sum = sum + row[i].value
    row[7].value = sum

# ##### Teacher sol.  #####
# scores = [
#     (1,10, 8, 5, 14, 26, 12), 
#     ......
# ]

# ws["H1"] = "총점"
# ws["I1"] = "성적"

# for idx, score in enumerate(scores, start=2):
#     sum_val = sum(score[1:]) - score[3] +10 
#     ws.cell(row=idx, column = 8).value = "=SUM(B{}:G{}".format(idx, idx)

#     grade = None
#     if sum_val >=90:
#         grade = "A"
#     elif sum_val >=80:
#         grade = "B"
#     elif sum_val >=70:
#         grade = "C"
#     else:
#         grade = "D"
#     if score[1] < 5:
#         grade = "F"
    
#     ws.cell(row=idx, column).value = grade  # I 열에 성적 정보 추가 

# %% [markdown]
# ## 학점 계산

# %%
row_range = ws[2:ws.max_row]
for row in row_range:
    if row[1].value < 5:
        row[8].value = "F"
    elif row[7].value >= 90:
        row[8].value = "A"
    elif row[7].value >= 80:
        row[8].value = "B"
    elif row[7].value >=70:
        row[8].value = "C"
    else:
        row[8].value = "D"


# %%
wb.save("quiz_1.xlsx")


# %%



