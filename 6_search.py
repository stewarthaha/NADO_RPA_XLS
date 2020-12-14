from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")
ws = wb.active  # 활성화 된 sheet

# 영어 대회 참가할 학생 추리기 , 80점 이상 학생
for row in ws.iter_rows(min_row=2): # 첫째줄 번호, 영어, 수학  은 건너뜀. 
    if int(row[1].value) > 80:
        print(row[0].value, " 번 학생은 영어 천재")

# 알고 봤더니 영어 점수가 아니고 컴퓨터 점수 였다,  영어-> 컴퓨터로 바꾸기 
for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value =="영어":
            cell.value = "컴퓨터"

wb.save("sample_modified.xlsx")

for x in range (1, ws.max_row +1):
    for y in range(1,ws.max_column+1):
        print(ws.cell(row=x,column=y).value, end=" ")
    print()      

