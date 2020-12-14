from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import coordinate_from_string

wb = Workbook()
ws = wb.active  # 활성화 된 sheet

# 1줄씩 데이터 넣기 
ws.append(["번호","영어","수학"])
for i in range(1,11):       # 10개 데이터 넣기 
    ws.append([i,randint(0,100), randint(0,100)])

# #영어 점수만 가져오기 
# col_B=ws["B"]  # 영어 컬럼만 가져오기 
# col_C=ws["C"]
# print(col_B)

# for cell in col_B:
#     print(cell.value)

# col_range=ws["B:C"] # 영어, 수학 column 함께 가져오기 

# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

# ### 첫번째 row 가져오기 
# row_title = ws[1]
# for cell in row_title:
#     print(cell.value)

# ## 여러 row 가져오기 
# row_range = ws[2:6]         # 2 ~ 6 줄까지 6 포함(주의 )
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

#### 마지막 줄까지 데이터 불러오기  
# row_range = ws[2:ws.max_row]  # 2번째 줄부터 마지막 줄까지 
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end=" ")     # A1 B1 C1   A2 B2 C2 .... 좌표정보 출력
#         xy=coordinate_from_string(cell.coordinate)
#         # print(xy, end=(" "))            # ('A', 2) ('B', 2) ('C', 2) .... 튜플 형태로 반환 

#         print(xy[0], end="")               # A B C A B C ....
#         print(xy[1], end=" ")                # 2 2 2 3 3 3 4 4 4 ..... 
#                                             # print(cell.coordinate, end=" ") 와 같은 출력값 
#     print()

#### 전체 rows 
# print(tuple(ws.rows))         # 한 row 씩 가져와서 튜플로

### 전체 colums 
# print(tuple(ws.columns))        # 한열씩 가져와서 튜플로...

### 전체 rows & columns 
# for row in tuple(ws.rows):
#     print(row)                  # (<Cell Sheet.A1>, <Cell Sheet.B1>, <Cell Sheet.C1>)......
#     print(row[1].value)         # 0:번호, 1: 영어, 2: 수학
# for column in tuple(ws.columns):
#     print(column[0].value)        # 번호, 영어, 수학 

### 다른 방법 
# for row in ws.iter_rows():        # 전체 row
#     print(row[2].value)             # 수학 점수 열 가져옴.

# for column in ws.iter_cols():       # 전체 column
#     print(column[0].value)          # 번호, 영어, 수학 

### 부분 선택하기 
for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3):    # 항상 다 채워줄 필요없다, 기본 최소,최대값 들어감.
    # print(row)
    print(row[0].value, row[1].value)   # 영어, 수학 






wb.save("sample.xlsx")
