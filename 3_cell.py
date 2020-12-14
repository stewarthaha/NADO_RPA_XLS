from openpyxl import Workbook
from random import *

wb = Workbook()    # 새 워크북 생성 
ws = wb.active
ws.title = "NadoSheet"  # sheet 이름 변경

# A1 셀에 1 입력 
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3
ws["b1"] = 4
ws["b2"] = 5
ws["b3"] = 6

print(ws["A1"])         # A1 셀의 정보 출력   <Cell NadoSheet.A1>
print(ws["A1"].value)   # 셀값 출력 , 값이 없을땐 None  출력

print(ws.cell(column=2, row=1).value)   # ws["B1"].value

ws.cell(column=3, row=1, value=10)
ws.cell(column=3, row=2).value=20

for i in range(3):
    print(ws.cell(column=3, row=i+1).value)

# 반복문을 이용해서 랜덤 숫자 채우기 
index=1
for x in range (1,11):          # 10개 row
    for y in range (1,11):      # 10개 column
        # ws.cell(row=x, column=y, value = randint(0,100))    # 0 ~ 100 사이의 숫자 
        ws.cell(row=x, column=y, value=index)
        index +=1


wb.save("sample.xlsx")        # 저장하기 
